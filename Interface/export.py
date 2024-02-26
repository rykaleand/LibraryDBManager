import openpyxl
import pandas as pd
from PyQt6 import QtCore, QtGui, QtWidgets
from PyQt6.QtWidgets import QMessageBox
from DB_connection.dbconnect import connect_to_Database

from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics

from SQl_files.sqls import get_headers, print_table
class Ui_Export(object):
    def setupUi(self, Export, table):
        self.table =table

        Export.setObjectName("Export")
        Export.resize(243, 193)
        self.centralwidget = QtWidgets.QWidget(parent=Export)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(parent=self.centralwidget)
        self.label.setGeometry(QtCore.QRect(60, 30, 141, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(11)
        font.setBold(False)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.comboBox = QtWidgets.QComboBox(parent=self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(70, 70, 111, 31))
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItem("")
        self.comboBox.addItem("")
        self.pushButton = QtWidgets.QPushButton(parent=self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(90, 130, 75, 24))
        self.pushButton.setObjectName("pushButton")

        #ЭКСПОРТИРУЕМ
        self.pushButton.clicked.connect(self.export_table_in_format)

        Export.setCentralWidget(self.centralwidget)

        self.retranslateUi(Export)
        QtCore.QMetaObject.connectSlotsByName(Export)

    def retranslateUi(self, Export):
        _translate = QtCore.QCoreApplication.translate
        Export.setWindowTitle(_translate("Export", "Export"))
        self.label.setText(_translate("Export", "Выберите формат:"))
        self.comboBox.setItemText(0, _translate("Export", ".pdf"))
        self.comboBox.setItemText(1, _translate("Export", ".xlsx"))
        self.pushButton.setText(_translate("Export", "OK"))

    def export_table_in_format(self):
        format = self.comboBox.currentText()
        exported_table_name = self.table
        if format == ".pdf":
            pdf_filename = "exported_" +  exported_table_name + ".pdf"
            self.export_table_to_pdf(pdf_filename)
        if format == ".xlsx":
            xls_filename = "exported_" + exported_table_name + ".xlsx"
            self.export_table_to_xls(xls_filename)

    def export_table_to_pdf(self, pdf_filename):
        selected_table = self.table
        try:

            column_names = get_headers(selected_table)
            rows = print_table(selected_table)

            pdfmetrics.registerFont(
                TTFont('Arial', 'arial.ttf'))

            df = pd.DataFrame(rows)

            pdf = SimpleDocTemplate(pdf_filename, pagesize=letter)

            # Регистрация шрифта с поддержкой кириллицы
            pdfmetrics.registerFont(TTFont('ArialUnicode','ARIALUNI.TTF'))
            data = [column_names] + df.values.tolist()

            table = Table(data, repeatRows=1, style=[('FONTNAME', (0, 0), (-1, -1), 'ArialUnicode')])

            # Добавление таблицы в PDF
            pdf.build([table])

            QMessageBox.information(None, "Succesfull", f"Таблица экспортирована " + pdf_filename)
            self.centralwidget.window().close()
        except Exception as e:
            QMessageBox.warning(None, "Warning", f"Failed to load table data: {str(e)}")
            print(str(e))

    def export_table_to_xls(self, xls_filename):
        selected_table = self.table
        try:

            column_names = get_headers(selected_table)
            rows = print_table(selected_table)

            book = openpyxl.Workbook()
            sheet = book.active

            # Запись заголовков в строку 1
            for col, head in enumerate(column_names, start=1):
                cell = sheet.cell(row=1, column=col)
                cell.value = head

            # Запись данных в таблицу
            for i, row in enumerate(rows, start=2):
                for j, col_value in enumerate(row, start=1):
                    cell = sheet.cell(row=i, column=j)
                    cell.value = col_value

            book.save(xls_filename)

            QMessageBox.information(None, "Succesfull", f"Table is exported as " + xls_filename)
            self.centralwidget.window().close()
        except Exception as e:
            QMessageBox.warning(None, "Warning", f"Failed to load table data: {str(e)}")
            print(str(e))
